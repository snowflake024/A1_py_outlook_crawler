from os import listdir
from list_of_services import list_of_services
import openpyxl as xl
import math

# Avoid providing a certain name of the topaz report for simplicity
list_of_files = listdir(r"C:\Users\a1bg482511\Desktop\lib\Repo\PYTHON\email_parser\attachments")
filename = list_of_files[0]

# Paths to various involved files
path_to_KPI = r"C:\Users\a1bg482511\Desktop\lib\Repo\PYTHON\email_parser\KPI_test\KPI_REPORT-v2.0.xlsx"
path_to_Topaz = rf"C:\Users\a1bg482511\Desktop\lib\Repo\PYTHON\email_parser\attachments\{filename}"
path_to_services = r"C:\Users\a1bg482511\Desktop\lib\Repo\PYTHON\email_parser\list_of_services.txt"

# Removed the password protection of the KPI workbook because of the added complexity to read it
kpi_wb = xl.load_workbook(path_to_KPI)
topaz_wb = xl.load_workbook(path_to_Topaz, read_only=True)
kpi_sheets = kpi_wb.sheetnames
topaz_sheet = topaz_wb["Verfügbarkeit pro Woche(in %)"]


def kpi_ws_months():
    """Get months in the form of a list for later use"""
    i = 1
    months = []
    while i < 13:
        months.append(kpi_sheets[i])
        i += 1
    return months


def round_up(n, decimals=0):
    """Rounding up to the nearest Hundredths place """
    """https://realpython.com/python-rounding/#pythons-built-in-round-function"""
    multiplier = 10 ** decimals
    return math.ceil(n * multiplier) / multiplier


def current_week_v1():
    """[TOPAZ]Check if the next cell in the KW* section is empty and return the previous one as current week"""
    # 5 + 52 column wise according to the topaz workbook
    week = 5
    while week < 57:
        if str(topaz_sheet.cell(2, week).value) == "None":
            week -= 1
            # it's better to return numerical value of column than cell contents
            # return topaz_sheet.cell(1, week).value
            return week
        else:
            week += 1


def get_service_data(kw):
    """Fetches the data for the current week"""

    # Create a dict with current week as first couple
    kpi_values = {'KW': topaz_sheet.cell(1, kw).value}
    week_column = kw

    for service in list_of_services:
        i = 0

        # Go through all rows in the first column to compare entries with service list
        for row in topaz_sheet.iter_rows():
            i += 1
            rows = f"A{i}"

            if service == topaz_sheet[rows].value:
                # print(topaz_sheet[rows].value)
                true_row = topaz_sheet[rows].row

                # Cell Coordinates in integers
                # print(f"{topaz_sheet[rows].row} {topaz_sheet[rows].column}")

                # Append all the services with their KPI % to the dict
                kpi_values[service] = round_up(topaz_sheet.cell(true_row, week_column).value, 2)
                break
    return kpi_values


def current_week_v2(topaz_data, kpi_months):
    """[KPI] Get current week based on the v1 of the same function but in the KPI wb"""

    # Get the current week that matches the topaz KW
    for sheet in kpi_months:
        topaz_week = f"W{topaz_data['KW'][2:4]}"
        col = 1
        while True:
            kpi_week = f"{kpi_wb[sheet].cell(1, col).value.partition(' ')[0]}"
            col += 1
            if kpi_week == "Monthly":
                break
            else:
                if kpi_week == topaz_week:
                    return [sheet, col]


def import_service_data(topaz_data, current_week, kpi_months):
    """"Fills in collected data"""

    # Set month and week values from v2 function
    week_column = current_week[1] - 1
    current_sheet = kpi_wb[current_week[0]]

    for service in list_of_services:
        i = 0
        # Go through all rows in the first column to compare entries with service list
        for row in current_sheet.iter_rows():
            i += 1
            rows = f"A{i}"

            if service == current_sheet[rows].value:
                service_row = current_sheet[rows].row

                # reading cell value from dict
                kpis = topaz_data[service]

                # writing the read value to destination excel file
                current_sheet.cell(row=service_row, column=week_column).value = kpis
                print("DONE")

    # saving the destination excel file
    kpi_wb.save(r"KPI_test\KPI_REPORT-v2.1.xlsx")

